from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries
import os
import locale
import shutil
import pandas as pd
from pandas import DataFrame
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side

from src.core.gateways.i_excel_service import IExcelService


class ExcelService(IExcelService):
    def __init__(self, caminho_arquivo):
        self.caminho_arquivo = caminho_arquivo
        self.workbook = self._get_workbook()

    def _get_workbook(self):
        """
        Abre um arquivo Excel existente ou cria um novo.
        Lança uma exceção clara se o arquivo estiver aberto (PermissionError).
        """
        try:
            if not os.path.exists(self.caminho_arquivo):
                # Tenta criar e salvar o novo workbook
                wb = Workbook()
                wb.save(self.caminho_arquivo)
            
            # Tenta carregar o workbook
            return load_workbook(self.caminho_arquivo)

        except PermissionError:
            # Captura o erro específico e lança uma nova exceção com uma mensagem amigável
            raise Exception(
                f"Permissão negada ao acessar o arquivo: '{self.caminho_arquivo}'.\n"
                f"Verifique se a planilha não está aberta em outro programa (como o Excel) e tente novamente."
            )

    def get_sheet(
        self, sheet_name: str, as_dataframe=False, header_row=1, columns=None
    ):
        """
        Retorna uma aba específica do arquivo Excel.
        Se as_dataframe=True, retorna como pandas DataFrame.
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Aba '{sheet_name}' não encontrada.")
        sheet = self.workbook[sheet_name]
        if not as_dataframe:
            return sheet

        # Extrai os dados da planilha
        data = list(sheet.values)
        if not data:
            return DataFrame()
        headers = data[header_row - 1]
        rows = data[header_row:]
        return DataFrame(rows, columns=headers)

    def delete_sheet(self, sheet_name):
        """
        Deleta uma planilha do arquivo Excel.

        Args:
            sheet_name (str): O nome da planilha a ser deletada.
        """
        if sheet_name in self.workbook.sheetnames:
            sheet_to_delete = self.workbook[sheet_name]
            self.workbook.remove(sheet_to_delete)

    def move_to_first(self, sheet_name: str):
        """
        Move a specific sheet to the first position in the Excel file.
        """
        try:
            # Find the sheet object
            sheet_to_move = self.workbook[sheet_name]

            # Remove the sheet from its current position in the internal list
            self.workbook._sheets.remove(sheet_to_move)

            # Insert the sheet at the beginning of the internal list
            self.workbook._sheets.insert(0, sheet_to_move)

        except KeyError:
            # Handle the case where the sheet isn't found
            raise ValueError(f"Sheet '{sheet_name}' not found.")

    def exportar_para_planilha(
        self,
        table: DataFrame,
        sheet_name: str,
        start_column="A",
        start_line="1",
        clear=False,
        sum_numeric=False,
        fit_columns=True,
        write_headers=True,
    ):

        # Cria aba se não existir
        if sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
        else:
            sheet = self.workbook.create_sheet(sheet_name)

        # Limpa a aba se necessário
        if clear:
            sheet.delete_rows(1, sheet.max_row)

        # Converte coluna e linha iniciais
        col_idx = column_index_from_string(start_column)
        row_idx = int(start_line)

        # Escreve cabeçalhos
        thin_border = Side(border_style="thin", color="000000")

        for j, column_name in enumerate(table.columns):
            cell = sheet.cell(row=row_idx, column=col_idx + j, value=column_name)
            cell.fill = PatternFill(
                start_color="4F81BD",
                end_color="4F81BD",
                fill_type="solid",
            )
            cell.border = Border(
                left=thin_border, right=thin_border, top=thin_border, bottom=thin_border
            )
            cell.font = Font(bold=True, color="FFFFFF")

        # Escreve os dados
        for row_offset, row in enumerate(table.itertuples(index=False)):
            border = Border(
                left=thin_border, right=thin_border, top=thin_border, bottom=thin_border
            )
            fill = PatternFill(
                start_color="DDEBF7" if row_offset % 2 == 0 else "FFFFFF",
                end_color="DDEBF7" if row_offset % 2 == 0 else "FFFFFF",
                fill_type="solid",
            )
            for col_offset, value in enumerate(row):
                cell = sheet.cell(
                    row=row_idx + 1 + row_offset,
                    column=col_idx + col_offset,
                    value=value,
                )

                cell.fill = fill
                cell.border = border

                col_name = table.columns[col_offset]
                if col_name.upper() in {
                    "VALOR",
                    "TOTAL",
                    "SALDO",
                    "DESCONTO",
                    "PROVENTO",
                } and isinstance(value, (int, float)):
                    cell.number_format = (
                        '_-R$ * #,##0.00_-;-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-'
                    )

        # Soma colunas numéricas (opcional)
        if sum_numeric:
            last_row = row_idx + len(table)
            for j, column in enumerate(table.columns):
                if table[column].dtype.kind in "iuf":  # int, unsigned, float
                    soma = table[column].sum()
                    sheet.cell(row=last_row + 1, column=col_idx + j, value=soma)

        # Fit Columns
        if fit_columns:
            self.fit_columns(sheet)

        if not write_headers:
            sheet.delete_rows(row_idx, 1)

        # Salva alterações
        self.save()

    def fit_columns(self, sheet):
        """Ajusta a largura das colunas da planilha."""
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column
            column_letter = get_column_letter(column)

            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = max_length + 7
            sheet.column_dimensions[column_letter].width = adjusted_width

    def delete_rows(self, sheet_name: str, start_row: int = 1):
        """
        Deleta todas as linhas em uma planilha a partir de uma linha específica.

        Args:
            sheet_name (str): O nome da aba.
            start_row (int): O número da linha a partir da qual as linhas serão deletadas.
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Aba '{sheet_name}' não encontrada.")

        sheet = self.workbook[sheet_name]
        num_rows_to_delete = sheet.max_row - start_row + 1

        if num_rows_to_delete > 0:
            sheet.delete_rows(start_row, num_rows_to_delete)
            self.save()

    # def apply_formula(self, sheet_name: str, formula: str, range_ref: str):
    #     """
    #     Aplica uma fórmula a um intervalo específico de células.
    #     """
    #     if sheet_name not in self.workbook.sheetnames:
    #         raise ValueError(f"Aba '{sheet_name}' não encontrada.")

    #     sheet = self.workbook[sheet_name]

    #     min_col, min_row, max_col, max_row = range_boundaries(range_ref)

    #     for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=min_col):
    #         for cell in row:
    #             current_row = cell.row
    #             # Formata a fórmula com a linha atual
    #             formatted_formula = formula.format(row=current_row)
    #             cell.value = formatted_formula
    #     self.save()

    def apply_number_format(self, sheet_name: str, number_format: str, range_ref: str):
        """
        Aplica um formato de número a um intervalo específico de células.

        Args:
            sheet_name (str): O nome da aba.
            number_format (str): O formato de número a ser aplicado (e.g., '#,##0.00').
            range_ref (str): O intervalo de células como uma string (e.g., 'A1:B10').
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Aba '{sheet_name}' não encontrada.")

        sheet = self.workbook[sheet_name]

        # O openpyxl.utils.range_boundaries retorna (min_col, min_row, max_col, max_row)
        min_col, min_row, max_col, max_row = range_boundaries(range_ref)

        for row in sheet.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
        ):
            for cell in row:
                cell.number_format = number_format

        self.save()

    def destacar_linhas(
        self,
        sheet_name: str,
        cor_fundo="FFFF00",
        negrito=False,
        coluna_alvo: str = None,
        valor_alvo=None,
        header_row=1,
    ):

        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Aba '{sheet_name}' não encontrada.")

        sheet = self.workbook[sheet_name]
        headers = [cell.value for cell in sheet[header_row]]

        try:
            idx_nat = headers.index(coluna_alvo)
        except ValueError:
            raise ValueError(f"Coluna '{coluna_alvo}' não encontrada na planilha.")

        fill = PatternFill(
            start_color=cor_fundo,
            end_color=cor_fundo,
            fill_type="solid",
        )
        font = Font(bold=negrito)
        thin_border = Side(border_style="thin", color="000000")
        border = Border(
            left=thin_border, right=thin_border, top=thin_border, bottom=thin_border
        )

        for row in sheet.iter_rows(min_row=header_row + 1):
            valor = row[idx_nat].value
            if str(valor) == valor_alvo:
                for cell in row:
                    cell.fill = fill
                    cell.border = border
                    cell.font = font

        self.save()

    def save(self):
        """Salva o arquivo Excel."""
        self.workbook.save(self.caminho_arquivo)

    @staticmethod
    def copy_to(caminho_arquivo: str, pasta_destino: str):
        """
        Copia o arquivo Excel para a pasta de destino especificada.
        Este método é estático e pode ser chamado sem instanciar a classe.

        Args:
            caminho_arquivo (str): O caminho completo do arquivo a ser copiado.
            pasta_destino (str): O caminho completo da pasta de destino.
        """
        # Garante que a pasta de destino exista
        if not os.path.isdir(pasta_destino):
            raise FileNotFoundError(
                f"A pasta de destino '{pasta_destino}' não foi encontrada."
            )

        # Constrói o caminho completo do novo arquivo na pasta de destino
        nome_arquivo = os.path.basename(caminho_arquivo)
        caminho_destino = os.path.join(pasta_destino, nome_arquivo)

        # Usa shutil.copy para copiar o arquivo
        shutil.copy(caminho_arquivo, caminho_destino)
        print(f"Arquivo copiado para: {caminho_destino}")

    @staticmethod
    def copy_data_with_pandas(caminho_origem: str, caminho_destino: str):
        """
        Copia os dados de um arquivo Excel para outro usando o pandas.
        """
        try:
            # Carrega todos os dados do arquivo de origem
            xlsx = pd.ExcelFile(caminho_origem)
            with pd.ExcelWriter(caminho_destino) as writer:
                # Itera sobre cada aba e a copia
                for sheet_name in xlsx.sheet_names:
                    df = pd.read_excel(xlsx, sheet_name)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

            print(f"Dados copiados de '{caminho_origem}' para '{caminho_destino}'.")

        except Exception as e:
            print(f"Ocorreu um erro: {e}")
