from datetime import datetime
import os
import sys
from openpyxl.styles import PatternFill, Font
import numpy as np
import pandas as pd
from pandas import DataFrame

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from drivers.pagamento.gerar_folha_pagamento import FolhaPagamento


ANO_ATUAL = datetime.now().year
MES_ATUAL = datetime.now().month
MESES = {
    1: "01-JANEIRO",
    2: "02-FEVEREIRO",
    3: "03-MARÇO",
    4: "04-ABRIL",
    5: "05-MAIO",
    6: "06-JUNHO",
    7: "07-JULHO",
    8: "08-AGOSTO",
    9: "09-SETEMBRO",
    10: "10-OUTUBRO",
    11: "11-NOVEMBRO",
    12: "12-DEZEMBRO",
}


class GerarConferencia:
    def __init__(self, nome_fundo: str):
        username = os.getlogin().strip()
        # Caminhos possíveis
        caminho_base = f"C:\\Users\\{username}\\Tribunal de Contas do Distrito Federal\\"
        caminho_onedrive = f"C:\\Users\\{username}\\OneDrive - Tribunal de Contas do Distrito Federal\\"

        # Escolher o primeiro caminho válido
        if os.path.exists(caminho_base):
            self.caminho_raiz = caminho_base
        elif os.path.exists(caminho_onedrive):
            self.caminho_raiz = caminho_onedrive
        else:
            raise FileNotFoundError(
                "Arquivo TEMPLATES_NL_{fundo}.xlsx não encontrado em nenhum dos caminhos possíveis.")

        self.nome_template = "CONFERÊNCIA"
        self.nome_fundo = nome_fundo.upper()

    def exportar_para_planilha(
        self,
        table: DataFrame,
        sheet_name: str,
        start_column="A",
        start_line="1",
        clear=False,
        sum_numeric=False,
    ):
        # Caminho do arquivo

        caminho_arquivo = (
            self.caminho_raiz
            + f"SECON - General\\ANO_ATUAL\\FOLHA_DE_PAGAMENTO_{ANO_ATUAL}\\{MESES[MES_ATUAL]}\\CONFERÊNCIA_{self.nome_fundo}.xlsx"
        )

        # Abre o arquivo existente
        if not os.path.exists(caminho_arquivo):
            # Cria um novo arquivo Excel
            wb = Workbook()
            wb.create_sheet(sheet_name)
            del wb["Sheet"]
            wb.save(caminho_arquivo)

        workbook = load_workbook(caminho_arquivo)

        # Cria aba se não existir
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)

        # Limpa a aba se necessário
        if clear:
            sheet.delete_rows(1, sheet.max_row)

        # Converte coluna e linha iniciais
        col_idx = column_index_from_string(start_column)
        row_idx = int(start_line)

        # Escreve cabeçalhos
        for j, column_name in enumerate(table.columns):
            sheet.cell(row=row_idx, column=col_idx + j, value=column_name)

        # Escreve os dados
        for row_offset, row in enumerate(table.itertuples(index=False)):
            for col_offset, value in enumerate(row):
                cell = sheet.cell(
                    row=row_idx + 1 + row_offset,
                    column=col_idx + col_offset,
                    value=value,
                )
                col_name = table.columns[col_offset]
                if col_name.upper() in {
                    "VALOR",
                    "TOTAL",
                    "SALDO",
                    "DESCONTO",
                    "PROVENTO",
                } and isinstance(value, (int, float)):
                    cell.number_format = (
                        '_-R$ * #,##0.00_-;-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-'
                    )

        # Soma colunas numéricas (opcional)
        if sum_numeric:
            last_row = row_idx + len(table)
            for j, column in enumerate(table.columns):
                if table[column].dtype.kind in "iuf":  # int, unsigned, float
                    soma = table[column].sum()
                    sheet.cell(row=last_row + 1,
                               column=col_idx + j, value=soma)

        # Fit Columns
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column  # Get the column index (number)
            column_letter = get_column_letter(column)  # Convert to letter

            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            # Set column width (add a little padding)
            adjusted_width = max_length + 5
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Salva alterações
        workbook.save(caminho_arquivo)

    def exportar_conferencia(self):
        folha_pagamento = FolhaPagamento(
            self.nome_fundo.lower(), "PRINCIPAL", test=True)

        dados_conferencia_financeiro = folha_pagamento.gerar_conferencia()

        proventos_folha = folha_pagamento.gerar_proventos(
            dados_conferencia_financeiro)
        descontos_folha = folha_pagamento.gerar_descontos(
            dados_conferencia_financeiro)

        # Exporta proventos na coluna A
        self.exportar_para_planilha(
            proventos_folha, self.nome_template, clear=True)

        # Exporta descontos na coluna G
        self.exportar_para_planilha(
            descontos_folha, self.nome_template, start_column="H", clear=False
        )

        # Calcula os totais
        total_proventos = (
            proventos_folha["PROVENTO"].sum(
            ) + descontos_folha["PROVENTO"].sum()
        )
        total_descontos = (
            proventos_folha["DESCONTO"].sum(
            ) + descontos_folha["DESCONTO"].sum()
        )
        total_liquido = total_proventos - total_descontos
        totais = DataFrame(
            {
                "TOTAIS": ["PROVENTOS", "DESCONTOS", "LÍQUIDO_FINANCEIRO"],
                "VALOR": [total_proventos, total_descontos, total_liquido],
            }
        )

        ultima_linha = len(descontos_folha) + 3

        # Exporta os totais abaixo dos dados
        self.exportar_para_planilha(
            totais,
            self.nome_template,
            start_column="H",
            start_line=str(ultima_linha),
            clear=False,
        )

    def exportar_adiantamento_ferias(self):
        folha_pagamento = FolhaPagamento(
            self.nome_fundo, "PRINCIPAL", test=True)
        df = folha_pagamento.gerar_conferencia(agrupar=False)

        df = df[df["AJUSTE"] == "ADIANTAMENTO FÉRIAS"]
        self.exportar_para_planilha(df, "ADIANTAMENTO_FÉRIAS")

    def exportar_nls(self):
        username = os.getlogin().strip()

        caminho_raiz = self.caminho_raiz + f"SECON - General\\CÓDIGOS\\"

        excel_rgps = pd.ExcelFile(
            caminho_raiz + "TEMPLATES_NL_RGPS.xlsx")
        templates_rgps = excel_rgps.sheet_names

        excel_financeiro = pd.ExcelFile(
            caminho_raiz + "TEMPLATES_NL_FINANCEIRO.xlsx")
        templates_financeiro = excel_financeiro.sheet_names

        excel_capitalizado = pd.ExcelFile(
            caminho_raiz + "TEMPLATES_NL_CAPITALIZADO.xlsx")
        templates_capitalizado = excel_capitalizado.sheet_names

        nomes_templates = {
            "RGPS": templates_rgps,
            "FINANCEIRO": templates_financeiro,
            "CAPITALIZADO": templates_capitalizado
        }

        # print(nomes_templates)

        drivers_pagamento = [
            FolhaPagamento(self.nome_fundo.lower(), nome_template, test=True) for nome_template in nomes_templates[self.nome_fundo]
        ]

        nls = {
            driver.nome_template: driver.gerar_folha() for driver in drivers_pagamento
        }

        print(nls.keys())
        for sheet_name, table_data in nls.items():
            print(f"Exportando {sheet_name}...")
            # print(table_data.head())
            self.exportar_para_planilha(table_data, sheet_name)


    def destacar_linhas(
        self,
        sheet_name: str,
        cor_fundo="FFFF00",
        negrito=False,
    ):
        """
        Destaca as linhas cuja coluna NME_NAT_DESPESA == 31901157
        """

        # Caminho do arquivo
        caminho_arquivo = (
            self.caminho_raiz
            + f"SECON - General\\ANO_ATUAL\\FOLHA_DE_PAGAMENTO_{ANO_ATUAL}\\{MESES[MES_ATUAL]}\\CONFERÊNCIA_{self.nome_fundo}.xlsx"
        )

        workbook = load_workbook(caminho_arquivo)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Aba '{sheet_name}' não encontrada.")

        sheet = workbook[sheet_name]

        # Cabeçalho
        headers = [cell.value for cell in sheet[1]]

        # Procura índice da coluna alvo
        try:
            idx_nat = headers.index("CDG_NAT_DESPESA")
        except ValueError:
            raise ValueError(
                "Coluna 'CDG_NAT_DESPESA' não encontrada na planilha.")

        # Estilo
        fill = PatternFill(start_color=cor_fundo,
                        end_color=cor_fundo, fill_type="solid")
        font = Font(bold=negrito)

        # Percorre linhas (pulando cabeçalho)
        for row in sheet.iter_rows(min_row=2):
            valor = row[idx_nat].value
            if str(valor) == "31901157":
                for i,cell in enumerate(row):
                    if i == 6:
                        break
                    cell.fill = fill
                    if negrito:
                        cell.font = font

        workbook.save(caminho_arquivo)

    def executar(self):
        self.exportar_conferencia()
        self.exportar_nls()

        if self.nome_fundo == "RGPS":
            self.exportar_adiantamento_ferias()


        self.destacar_linhas(
            sheet_name="CONFERÊNCIA",
            cor_fundo="FF9999",  # vermelho claro
            negrito=True
        )


if __name__ == "__main__":
    try:
        for fundo in ["CAPITALIZADO"]:
            gerador = GerarConferencia(fundo)
            gerador.executar()
    except Exception as e:
        print(e)
        sys.exit()
