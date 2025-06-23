from datetime import datetime
import os
import sys

import numpy as np
import pandas as pd
from pandas import DataFrame

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from drivers.rgps.rgps_base import (
    PreenchimentoRGPSBeneficios,
    PreenchimentoRGPSDeaBeneficios,
    PreenchimentoRGPSIndenizacoesRestituicoes,
    PreenchimentoRGPSPrincipal,
    PreenchimentoRGPSSubstituicoes,
    PreenchimentoRGPSIndenizacoesPessoal,
)

ANO_ATUAL = datetime.now().year
MES_ATUAL = datetime.now().month
MESES = {
    1: "01-JANEIRO",
    2: "02-FEVEREIRO",
    3: "03-MARÇO",
    4: "04-ABRIL",
    5: "05-MAIO",
    6: "06-JUNHO",
    7: "07-JULHO",
    8: "08-AGOSTO",
    9: "09-SETEMBRO",
    10: "10-OUTUBRO",
    11: "11-NOVEMBRO",
    12: "12-DEZEMBRO",
}


class GerarConferencia:
    def __init__(self):
        username = os.getlogin().strip()
        self.caminho_raiz = f"C:\\Users\\{username}\\Tribunal de Contas do Distrito Federal\\"
        self.nome_template = "CONFERÊNCIA"

    def exportar_para_planilha(
        self,
        table: DataFrame,
        sheet_name: str,
        start_column="A",
        start_line="1",
        clear=False,
        sum_numeric=False,
    ):
        # Caminho do arquivo
        # caminho_arquivo = (
        #     self.caminho_raiz
        #     + f"SECON - General\\CÓDIGOS\\TEMPLATES_NL_RGPS - Copia.xlsx"
        # )
        caminho_arquivo = (
            self.caminho_raiz
            + f"SECON - General\\ANO_ATUAL\\FOLHA_DE_PAGAMENTO_{ANO_ATUAL}\\{MESES[MES_ATUAL]}\\CONFERÊNCIA_RGPS.xlsx"
        )

        # Abre o arquivo existente
        if not os.path.exists(caminho_arquivo):
            # Cria um novo arquivo Excel
            wb = Workbook()
            wb.create_sheet(sheet_name)
            del wb["Sheet"]
            wb.save(caminho_arquivo)

        workbook = load_workbook(caminho_arquivo)

        # Cria aba se não existir
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)

        # Limpa a aba se necessário
        if clear:
            sheet.delete_rows(1, sheet.max_row)

        # Converte coluna e linha iniciais
        col_idx = column_index_from_string(start_column)
        row_idx = int(start_line)

        # Escreve cabeçalhos
        for j, column_name in enumerate(table.columns):
            sheet.cell(row=row_idx, column=col_idx + j, value=column_name)

        # Escreve os dados
        for row_offset, row in enumerate(table.itertuples(index=False)):
            for col_offset, value in enumerate(row):
                cell = sheet.cell(
                    row=row_idx + 1 + row_offset,
                    column=col_idx + col_offset,
                    value=value,
                )
                col_name = table.columns[col_offset]
                if col_name.upper() in {
                    "VALOR",
                    "TOTAL",
                    "SALDO",
                    "DESCONTO",
                    "PROVENTO",
                } and isinstance(value, (int, float)):
                    cell.number_format = (
                        '_-R$ * #,##0.00_-;-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-'
                    )

        # Soma colunas numéricas (opcional)
        if sum_numeric:
            last_row = row_idx + len(table)
            for j, column in enumerate(table.columns):
                if table[column].dtype.kind in "iuf":  # int, unsigned, float
                    soma = table[column].sum()
                    sheet.cell(row=last_row + 1, column=col_idx + j, value=soma)

        # Fit Columns
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column  # Get the column index (number)
            column_letter = get_column_letter(column)  # Convert to letter

            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            # Set column width (add a little padding)
            adjusted_width = max_length + 5
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Salva alterações
        workbook.save(caminho_arquivo)

    def gerar_conferencia(self,agrupar=True):
        def cria_coluna_rubrica(row):  # lógica redundante?
            cdg_nat_despesa = str(row["CDG_NAT_DESPESA"])
            valor = row["VALOR_AUXILIAR"]

            if cdg_nat_despesa.startswith("3"):
                if valor > 0:
                    return "PROVENTO"
                else:
                    return "DESCONTO"
            elif cdg_nat_despesa.startswith("2"):
                if valor < 0:
                    return "DESCONTO"
                else:
                    return "PROVENTO"
            else:
                return ""  # Ou algum outro valor padrão, se necessário

        def categorizar_rubrica(rubrica):
            if rubrica in [11962, 21962, 32191, 42191, 52191, 61962]:
                return "ADIANTAMENTO FÉRIAS"
            return ""

        caminho_planilha = f"SECON - General\\ANO_ATUAL\\FOLHA_DE_PAGAMENTO_{ANO_ATUAL}\\{MESES[MES_ATUAL]}\\DEMOFIN_TABELA.xlsx"
        # caminho_planilha = f"SECON - General\\ANO_ATUAL\\FOLHA_DE_PAGAMENTO_2025\\03-MARÇO\\DEMOFIN_TABELA.xlsx"

        caminho_completo = self.caminho_raiz + caminho_planilha

        plan_folha = pd.read_excel(caminho_completo, sheet_name="DEMOFIN - T", header=1)

        plan_folha["CDG_NAT_DESPESA"] = plan_folha["CDG_NAT_DESPESA"].str.replace(
            ".", ""
        )

        filtro_fundo = plan_folha["CDG_FUNDO"] == 1
        plan_rgps = plan_folha.loc[
            filtro_fundo,
            ["CDG_PROVDESC", "NME_NAT_DESPESA", "CDG_NAT_DESPESA", "VALOR_AUXILIAR"],
        ]

        plan_rgps["RUBRICA"] = plan_rgps.apply(cria_coluna_rubrica, axis=1)

        plan_rgps["VALOR_AUXILIAR"] = plan_rgps["VALOR_AUXILIAR"].abs()

        plan_rgps = pd.pivot_table(
            plan_rgps,
            values="VALOR_AUXILIAR",
            index=["CDG_PROVDESC", "NME_NAT_DESPESA", "CDG_NAT_DESPESA"],
            columns="RUBRICA",
            aggfunc="sum",
        )

        conferencia_rgps = (
            plan_rgps.groupby(["CDG_PROVDESC", "NME_NAT_DESPESA", "CDG_NAT_DESPESA"])[
                ["PROVENTO", "DESCONTO"]
            ]
            .agg(lambda x: np.sum(np.abs(x)))
            .reset_index()
        )

        # Aplica a função pra criar a coluna "AJUSTE"
        conferencia_rgps["AJUSTE"] = conferencia_rgps["CDG_PROVDESC"].apply(
            categorizar_rubrica
        )
        conferencia_rgps.sort_values(by=["CDG_NAT_DESPESA"])

        if not agrupar:
            return conferencia_rgps

        conferencia_rgps_final = (
            conferencia_rgps.groupby(["NME_NAT_DESPESA", "CDG_NAT_DESPESA"])[
                ["PROVENTO", "DESCONTO"]
            ]
            .agg(lambda x: np.sum(np.abs(x)))
            .reset_index()
            .sort_values(by="CDG_NAT_DESPESA", ascending=True)
        )

        return conferencia_rgps_final

    def gerar_proventos(self, conferencia_rgps_final: DataFrame):
        prov_rgps = conferencia_rgps_final.loc[
            conferencia_rgps_final["CDG_NAT_DESPESA"].str.startswith("3")
        ]

        coluna_saldo_proventos = prov_rgps["PROVENTO"] - prov_rgps["DESCONTO"]
        prov_rgps_final = prov_rgps.loc[
            :, ["NME_NAT_DESPESA", "CDG_NAT_DESPESA", "PROVENTO", "DESCONTO"]
        ].assign(SALDO=coluna_saldo_proventos)

        prov_rgps_final["CDG_NAT_DESPESA"] = prov_rgps_final[
            "CDG_NAT_DESPESA"
        ].str.slice(1)

        prov_rgps_final.sort_values(by=["CDG_NAT_DESPESA"])

        return prov_rgps_final

    def gerar_descontos(self, conferencia_rgps_final: DataFrame):
        desc_rgps = conferencia_rgps_final.loc[
            ~conferencia_rgps_final["CDG_NAT_DESPESA"].str.startswith("3")
        ]
        desc_rgps

        coluna_saldo_descontos = desc_rgps["DESCONTO"] - desc_rgps["PROVENTO"]
        coluna_saldo_descontos
        desc_rgps = desc_rgps.loc[
            :, ["NME_NAT_DESPESA", "CDG_NAT_DESPESA", "DESCONTO", "PROVENTO"]
        ].assign(SALDO=coluna_saldo_descontos)
        desc_rgps.sort_values(by=["CDG_NAT_DESPESA"])
        return desc_rgps

    def exportar_conferencia(self):
        dados_conferencia_rgps = self.gerar_conferencia()
        # Gera os dados
        proventos_rgps = self.gerar_proventos(dados_conferencia_rgps)
        descontos_rgps = self.gerar_descontos(dados_conferencia_rgps)

        # Exporta proventos na coluna A
        self.exportar_para_planilha(proventos_rgps, self.nome_template, clear=True)

        # Exporta descontos na coluna G
        self.exportar_para_planilha(
            descontos_rgps, self.nome_template, start_column="G", clear=False
        )

        # Calcula os totais
        total_proventos = (
            proventos_rgps["PROVENTO"].sum() + descontos_rgps["PROVENTO"].sum()
        )
        total_descontos = (
            proventos_rgps["DESCONTO"].sum() + descontos_rgps["DESCONTO"].sum()
        )
        total_liquido = total_proventos - total_descontos
        totais = DataFrame(
            {
                "TOTAIS": ["PROVENTOS", "DESCONTOS", "LÍQUIDO_RGPS"],
                "VALOR": [total_proventos, total_descontos, total_liquido],
            }
        )

        ultima_linha = len(descontos_rgps) + 3

        # Exporta os totais abaixo dos dados
        self.exportar_para_planilha(
            totais,
            self.nome_template,
            start_column="J",
            start_line=str(ultima_linha),
            clear=False,
        )

    def exportar_adiantamento_ferias(self):
        df = self.gerar_conferencia(False)
        df = df[df["AJUSTE"] == "ADIANTAMENTO FÉRIAS"]
        self.exportar_para_planilha(df, "ADIANTAMENTO_FÉRIAS")

    def exportar_nls(self):
        drivers_rgps = [
            PreenchimentoRGPSPrincipal(test=True),
            PreenchimentoRGPSSubstituicoes(test=True),
            PreenchimentoRGPSBeneficios(test=True),
            PreenchimentoRGPSIndenizacoesRestituicoes(test=True),
            PreenchimentoRGPSIndenizacoesPessoal(test=True),
            PreenchimentoRGPSDeaBeneficios(test=True),
        ]

        nls = {
            driver.nome_template: driver.gerar_folha_rgps() for driver in drivers_rgps
        }

        for sheet_name, table_data in nls.items():
            self.exportar_para_planilha(table_data, sheet_name)

    def executar(self):
        self.exportar_conferencia()
        self.exportar_nls()
        self.exportar_adiantamento_ferias()


try:
    gerador = GerarConferencia()
    gerador.executar()
except Exception as e:
    print(e)
    sys.exit()
