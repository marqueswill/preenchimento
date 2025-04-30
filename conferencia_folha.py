from datetime import datetime
import os
from openpyxl import load_workbook
import pandas as pd

ANO_ATUAL = datetime.now().year
MES_ATUAL = datetime.now().month
MESES = {
    1: "01-JANEIRO",
    2: "02-FEVEREIRO",
    3: "03-MARÇO",
    4: "04-ABRIL",
    5: "05-MAIO",
    6: "06-JUNHO",
    7: "07-JULHO",
    8: "08-AGOSTO",
    9: "09-SETEMBRO",
    10: "10-OUTUBRO",
    11: "11-NOVEMBRO",
    12: "12-DEZEMBRO",
}


class GerarConferencia:
    def __init__(self, test=False):
        self.test = test
        self.username = os.getlogin().strip()

    def gerar_conferencia_rgps(self):
        plan_folha = pd.read_excel(
            f"C:\\Users\\{self.username}\\OneDrive - Tribunal de Contas do Distrito Federal\\General - SECON\\ANO_ATUAL\\FOLHA_DE_PAGAMENTO_{ANO_ATUAL}\\{MESES[MES_ATUAL]}\\DEMOFIN_TABELA.xlsx",
            sheet_name="DEMOFIN - T",
            header=1,
        )

        plan_folha["CDG_NAT_DESPESA"] = plan_folha["CDG_NAT_DESPESA"].str.replace(
            ".", ""
        )

        filtro_fundo = plan_folha["NME_FUNDO"] == "RGPS"
        plan_rgps_copia = plan_folha.loc[
            filtro_fundo, ["CDG_NAT_DESPESA", "NME_PROVDESC", "VALOR", "RUBRICA"]
        ]

        # Função para criar a coluna "AJUSTE"
        def categorizar_rubrica(rubrica):
            if rubrica in [11962, 21962, 32191, 42191, 52191, 61962]:
                return "Adiantamento_ferias"
            return ""

        # Aplica a função pra criar a coluna "AJUSTE"
        plan_rgps_copia["AJUSTE"] = plan_folha["CDG_PROVDESC"].apply(
            categorizar_rubrica
        )
        plan_rgps_copia.reset_index(drop=True, inplace=True)

        # Calcula as somas dos valores para RUBRICA "Provento" e "Desconto"
        soma_provento = plan_rgps_copia.loc[
            plan_rgps_copia["RUBRICA"] == "Provento", "VALOR"
        ].sum()
        soma_desconto = plan_rgps_copia.loc[
            plan_rgps_copia["RUBRICA"] == "Desconto", "VALOR"
        ].sum()

        if self.test:
            caminho_planilha = f"C:\\Users\\{self.username}\\OneDrive - Tribunal de Contas do Distrito Federal\\General - SECON\\CÓDIGOS\\PLANILHA_FOLHA_NORMAL_RGPS_MARCO_2025_NOVA.xlsx"
        else:
            caminho_planilha = f"C:\\Users\\{self.username}\\OneDrive - Tribunal de Contas do Distrito Federal\\General - SECON\\ANO_ATUAL\\FOLHA_DE_PAGAMENTO_{ANO_ATUAL}\\{MESES[MES_ATUAL]}\\PLANILHA_FOLHA_NORMAL_RGPS_{MESES[MES_ATUAL].split('-')[1].replace('Ç','C')}_{ANO_ATUAL}.xlsx"

        # Carregar a planilha
        wb = load_workbook(caminho_planilha)

        if "CONFERENCIA_RGPS" in wb.sheetnames:
            ws = wb["CONFERENCIA_RGPS"]
            ws.delete_rows(1, ws.max_row)  # Limpa a aba (mantém cabeçalhos)

            headers = list(plan_rgps_copia.columns)
            for j, header in enumerate(headers, start=1):
                ws.cell(row=1, column=j, value=header)

            # Inserir novos dados manualmente, linha por linha
            for i, row in enumerate(plan_rgps_copia.itertuples(index=False), start=2):
                for j, value in enumerate(row, start=1):
                    ws.cell(row=i, column=j, value=value)

            # Pula uma linha após os dados
            ultima_linha = ws.max_row + 1

            # Adiciona as somas com rótulos
            ws.cell(row=ultima_linha + 1, column=2, value="Soma Proventos")
            ws.cell(row=ultima_linha + 1, column=3, value=soma_provento)

            ws.cell(row=ultima_linha + 2, column=2, value="Soma Descontos")
            ws.cell(row=ultima_linha + 2, column=3, value=soma_desconto)

            ws.cell(row=ultima_linha + 3, column=2, value="Líquido RGPS")
            ws.cell(row=ultima_linha + 3, column=3, value=soma_provento - soma_desconto)

            wb.save(caminho_planilha)


# dataframe = pd.read_excel(caminho_planilha, sheet_name="VALORES_RGPS")

gerador = GerarConferencia(test=True)
gerador.gerar_conferencia_rgps()
